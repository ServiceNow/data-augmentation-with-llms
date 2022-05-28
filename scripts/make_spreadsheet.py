import os, json, pickle
from openpyxl import Workbook
from openpyxl.styles import Font

DATASET_NAME = "banking77"


def load_data():
    data_dir = f"./data/{DATASET_NAME}"
    opj = os.path.join
    ds_full_suite = pickle.load(open(opj(data_dir, "full/data_full_suite.pkl"), "rb"))
    generated_samples = pickle.load(open(opj(data_dir, "full/al_dataset.pkl"), "rb"))[
        "generated"
    ]
    id2name = json.load(open(opj(data_dir, "id2name.json")))
    return ds_full_suite, generated_samples, id2name


def massage_data(ds_full_suite, generated_samples, id2name):
    workbooks = {}
    for engine in generated_samples:
        workbooks[engine] = gather_workbook_data(
            ds_full_suite, generated_samples[engine], id2name
        )
    return workbooks


def gather_workbook_data(ds_full_suite, generated_samples, id2name):
    workbook_data = {}
    for domain in ds_full_suite:
        # generate prompt column
        prompt_data = ds_full_suite[domain]["F"]["train"]
        for text, intent_id in zip(prompt_data["text"], prompt_data["intent"]):
            intent_name = id2name[str(intent_id)].replace("?", "")
            sheet_name = f"{domain}<>{intent_name}"
            if sheet_name not in workbook_data:
                workbook_data[sheet_name] = {
                    "prompt": [],
                    "generated": [],
                    "oracle_prediction": [],
                }
            workbook_data[sheet_name]["prompt"].append(text)

        # add generated data, and oracle prediction data
        for text, oracle_intent_id, org_intent_id in zip(
            generated_samples["text"],
            generated_samples["intent"],
            generated_samples["old_intent"],
        ):
            oracle_pred = id2name[str(oracle_intent_id)].replace("?", "")
            org_intent_name = id2name[str(org_intent_id)].replace("?", "")
            sheet_name = f"{domain}<>{org_intent_name}"
            if sheet_name not in workbook_data:
                # print(f"sheet {sheet_name} doesn't exist")
                continue
            workbook_data[sheet_name]["generated"].append(text)
            workbook_data[sheet_name]["oracle_prediction"].append(oracle_pred)
    return workbook_data


def create_excel_sheet(name, data):
    wb = Workbook()
    wb.remove(wb.active)  # remove the empty "Sheet"
    # create different sheets
    for sheet_name in data:
        org_intent = sheet_name.split("<>", 1)[1]
        ws = wb.create_sheet(sheet_name)
        prompts = data[sheet_name]["prompt"]
        generated = data[sheet_name]["generated"]
        oracle_predictions = data[sheet_name]["oracle_prediction"]

        ############# compute some quantities for formatting  ##############
        # max width of column A
        max_sent_length = max(map(len, prompts + generated))
        # max width of column B
        max_pred_length = max(map(len, oracle_predictions))
        total_faithful_samples = oracle_predictions.count(org_intent)
        ############# compute end #################

        # add the first column
        ws.append(["Sentences", "Oracle Predictions"])
        # add the sentences column
        for irow in range(len(prompts + generated)):
            if irow < len(prompts):
                ws.append([prompts[irow]])
            else:
                new_irow = irow - len(prompts)
                ws.append([generated[new_irow], oracle_predictions[new_irow]])

        # some analysis
        ws["C1"] = "Total faithful samples"
        ws["C2"] = f"{total_faithful_samples}/{len(generated)}"
        ws["C3"] = f"{total_faithful_samples/len(generated)*100:.2f}%"

        # adjust column widths
        ws.column_dimensions["A"].width = max_sent_length
        ws.column_dimensions["B"].width = max_pred_length
        ws.column_dimensions["C"].width = len("Total faithful samples")

        # increase font size
        n_rows = len(prompts + generated)
        for col, n_rows in [("A", n_rows), ("B", n_rows), ("C", 3)]:
            for i in range(1, n_rows + 2):
                ws[f"{col}{i}"].font = Font(size=14)

        # bold the first row
        ws["A1"].font = Font(bold=True, size=14)
        ws["B1"].font = Font(bold=True, size=14)
        ws["C1"].font = Font(bold=True, size=14)

    # delete this useless sheet
    # sort sheets based on fidelity (ws['C3'] is the fidelity)
    wb._sheets.sort(key=lambda ws: float(ws["C3"].value[:-1]))
    wb.active = 0

    save_folder = f"spreadsheets/{DATASET_NAME}"
    if not os.path.exists(save_folder):
        os.mkdir(save_folder)
    wb.save(os.path.join(save_folder, f"{name}.xlsx"))


if __name__ == "__main__":
    workbooks = massage_data(*load_data())
    for engine_temp, data in workbooks.items():
        if "_" in engine_temp and engine_temp.split("_")[1] != "1.0":
            continue
        create_excel_sheet(engine_temp, data)
