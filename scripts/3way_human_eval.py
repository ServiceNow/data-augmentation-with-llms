from collections import defaultdict
from utils import main_data_utils as mdu
import argparse, numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font


def create_content(ds, id2name, args, run_code):
    samples = []
    for (intent_id, sentence) in zip(ds[run_code]["intent"], ds[run_code]["text"]):
        if id2name[str(intent_id)] not in args.intent_triplet:
            continue
        samples.append((sentence, id2name[str(intent_id)]))

    np.random.shuffle(samples)
    # start creating an excel sheet
    wb = Workbook()
    wb.remove(wb.active)

    # create a fresh worksheet with meaningful name
    ws = wb.create_sheet("Human eval.")
    ws.append(["Sentence", "Your Prediction"])
    labels = ""
    max_sent_length = 0
    for sent, intent in samples:
        if len(sent) > max_sent_length:
            max_sent_length = len(sent)
        ws.append([sent])
        labels += intent + "\n"

    # increase font size of sentences
    for i in range(1, len(samples) + 2):
        ws[f"A{i}"].font = Font(size=14)

    # max width of column A
    ws.column_dimensions["A"].width = max_sent_length
    ws["A1"].font = Font(bold=True, size=14)
    ws["B1"].font = Font(bold=True, size=14)
    wb.active = 0
    wb.save(f"spreadsheets/{args.dataset_name}/{run_code}_human_eval.xlsx")
    mdu.write_file(labels, f"spreadsheets/{args.dataset_name}/{run_code}_labels.txt")


def evaluate(args, run_code):
    labels_path = f"spreadsheets/{args.dataset_name}/{run_code}_labels.txt"
    labels = mdu.read_file(labels_path).splitlines()
    if args.dataset_name == "hwu64":
        predcode2name = {
            "1": "music_likeness",
            "2": "music_settings",
            "3": "play_music",
        }
    elif args.dataset_name == "banking77":
        predcode2name = {
            "1": "topping_up_by_card",
            "2": "top_up_failed",
            "3": "pending_top_up",
        }
    preds_path = f"spreadsheets/{args.dataset_name}/sahu_{run_code}_preds.txt"
    preds = [predcode2name[p] for p in mdu.read_file(preds_path).splitlines()]
    assert len(preds) == len(labels)
    class_wise_preds = defaultdict(list)
    for _pred, _label in zip(preds, labels):
        class_wise_preds[_label].append(_pred)
    print(
        "Overall Val Acc of sahu =",
        f"{np.mean([1 if p == l else 0 for p, l in zip(preds, labels)])*100:.2f}",
    )

    for intent, preds in class_wise_preds.items():
        print(f"sahu Acc on {intent}: {preds.count(intent)/len(preds)*100:.2f}")


def main():
    # read base data
    args = parse_args()
    if args.eval:
        evaluate(args, "val")
    ds = mdu.read_pickle(f"data/{args.dataset_name}/full/dataset.pkl")
    id2name = mdu.read_json(f"data/{args.dataset_name}/id2name.json")

    create_content(ds, id2name, args, "val")  # 3-way val set
    create_content(ds, id2name, args, "test")  # 3-way test set


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("-d", "--dataset-name", default="hwu64")
    # parser.add_argument("-d", "--dataset-name", default="banking77")
    parser.add_argument(
        "-it",
        "--intent-triplet",
        nargs="+",
        default=["music_likeness", "play_music", "music_settings"],
        # default=["topping_up_by_card", "top_up_failed", "pending_top_up"],
    )
    parser.add_argument("--eval", action="store_true")
    return parser.parse_args()


if __name__ == "__main__":
    main()
